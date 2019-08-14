import mysql.connector as mariadb
from openpyxl import load_workbook
from collections import defaultdict

conn = mariadb.connect(user='root', database='mpce1')
cur = conn.cursor()

initial_agent_mapping = []

with open("C:/git/mpce-database-reform/hidden/deduplication_sheets/agents.xlsx", "rb") as pth:
    wbk = load_workbook(pth, read_only=True, keep_vba=False)
    initial_agent_mapping = [(row[0], row[7]) for row in wbk['pairwise_agent_comparison'].iter_rows(
        min_row=2, max_row=2865, max_col=17, values_only=True) if row[16] == 'Y']

agent_mapping_dict = {}
already_entered = set()
for (id_1, id_2) in initial_agent_mapping:
    if id_1 in already_entered:
        if id_2 in already_entered:
            continue
        elif id_1 in agent_mapping_dict:
            agent_mapping_dict[id_1].add(id_2)
        else:
            for key, values in agent_mapping_dict.items():
                if id_1 in values:
                    agent_mapping_dict[key].add(id_2)
    elif id_2 in already_entered:
        if id_2 in agent_mapping_dict:
            agent_mapping_dict[id_2].add(id_1)
        else:
            for key, values in agent_mapping_dict.items():
                if id_2 in values:
                    agent_mapping_dict[key].add(id_1)
    else:
        agent_mapping_dict[id_1] = {id_2}
    already_entered.add(id_1)
    already_entered.add(id_2)

final_agent_mapping = []
for key, values in agent_mapping_dict.items():
    for value in values:
        final_agent_mapping.append((key, value))

cur.execute('DROP TABLE IF EXISTS final_agent_mapping')

cur.execute('CREATE TABLE final_agent_mapping (id_1 CHAR(8), id_2 CHAR(8))')
cur.executemany('INSERT INTO final_agent_mapping VALUES (%s, %s)', seq_params=final_agent_mapping)

cur.execute('CREATE INDEX map ON final_agent_mapping (id_2, id_1)')

conn.commit()
cur.close()
conn.close()
