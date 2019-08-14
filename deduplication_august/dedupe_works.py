import mysql.connector as mariadb
from openpyxl import load_workbook
from collections import defaultdict

conn = mariadb.connect(user='root', database='mpce1')
cur = conn.cursor()

initial_mapping = []

with open("C:/git/mpce-database-reform/hidden/deduplication_sheets/work.xlsx", "rb") as pth:
    wbk = load_workbook(pth, read_only=True, keep_vba=False)
    initial_mapping = [(row[0], row[3], row[10] - 1) for row in wbk['pairwise_work_comparisons'].iter_rows(
        min_row=2, max_row=640, max_col=11, values_only=True) if row[8] == 'Y']

print(f'{len(initial_mapping)} duplicate work pairs found.')

mapping_dict = {}
already_entered = set()
for tup in initial_mapping:
    ids = tup[0:2]
    preferred = tup[2]
    if ids[preferred] in already_entered:
        if ids[preferred] in mapping_dict:
            mapping_dict[ids[preferred]].add(ids[1 - preferred])
        else:
            for key, values in mapping_dict.items():
                if ids[preferred] in values:
                    mapping_dict[key].add(ids[1 - preferred])
    else:
        mapping_dict[ids[preferred]] = {ids[1 - preferred]}
        
    already_entered.add(ids[0])
    already_entered.add(ids[1])

print(f'{len(already_entered)} works to be deduplicated.')

final_mapping = []
for key, values in mapping_dict.items():
    for value in values:
        final_mapping.append((key, value))

print(f'After reducing map, {len(final_mapping)} duplicate pairs found.')

cur.execute('DROP TABLE IF EXISTS final_work_mapping')

cur.execute('CREATE TABLE final_work_mapping (id_1 CHAR(12), id_2 CHAR(12))')
cur.executemany('INSERT INTO final_work_mapping VALUES (%s, %s)', seq_params=final_mapping)

cur.execute('CREATE INDEX map ON final_work_mapping (id_2, id_1)')

conn.commit()
cur.close()
conn.close()
