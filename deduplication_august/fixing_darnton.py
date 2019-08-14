# There was an error in the import from darnton.xlsx. I misinterpreted the first two columns.
# Let's fix it on the local version! Then upload some sql that will fix it on the remote.

import mysql.connector as mariadb
from openpyxl import load_workbook
import re

conn = mariadb.connect(user="root", db="mpce1")
cur = conn.cursor()

# Get works
cur.execute("SELECT * FROM work WHERE illegality_notes")
works = cur.fetchall()
print(f'{len(works)} works fetched from database, for example:\n{works[100]}')

# Get spreadsheet data
with open("C:/git/mpce-database-reform/mpcereform/spreadsheets/darnton.xlsx", 'rb') as pth:
    darnton = load_workbook(pth, read_only=True, keep_vba=False)
    all_darnton = [row for row in darnton['Sheet1'].iter_rows(min_row=2, max_row=175, max_col=19, values_only=True)]

print(f'{len(all_darnton)} works imported from spreadsheet, for example:\n{all_darnton[20]}')

# Create dict from works
darn_rgx = re.compile(r"\d+ \(\d+\) Darnton")
def extr(ill_note):
    out = darn_rgx.search(ill_note)[0]
    return out

works = {(work[1][:7], extr(work[3])):[work[0],work[1]] for work in works if darn_rgx.search(work[3])}

darnton_linked = [list(works[(darnton[3][:7], f'{darnton[0]} ({darnton[1]}) Darnton')]) + list(darnton) for darnton in all_darnton]

# Error checking
# Do all the books match?
non_matching = [rec for rec in darnton_linked if rec[1] != rec[5]]
print(f'{len(non_matching)} titles do not match:\n{non_matching}')

# Generating sql
updates = ','.join(
    [f"('{rec[0]}', '0 (0) Darnton')" for rec in darnton_linked])
stmt = f"""USE mpce1;
CREATE TEMPORARY TABLE works_to_update (work_code CHAR(12), new_illegality TEXT);
INSERT INTO works_to_update VALUES
{updates};
UPDATE work, works_to_update
SET work.illegality_notes = works_to_update.new_illegality
WHERE work.work_code = works_to_update.work_code;"""

with open('fix_darnton.sql', 'wt') as pth:
    pth.write(stmt)
    print(f'SQL written to fix_darnton.sql')
# Loop 

conn.close()

# SQL generated, run on server successfully 20190812