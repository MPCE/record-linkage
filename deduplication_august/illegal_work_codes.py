from openpyxl import load_workbook

with open("deduplication_sheets/bastille_codes.xlsx", "rb") as pth:
    bastille_wbk = load_workbook(pth, read_only=True, keep_vba=False)
    bastille_assigns = [
        (row[0], row[4]) for row in bastille_wbk['bastille_work_codes'].iter_rows(
            min_row=2, max_row=47, max_col=8, values_only=True 
        ) if row[7] == 'Y'
    ]

sql = f"""
    INSERT INTO bastille_register_record (ID, work_code)
    VALUES {str(bastille_assigns)[1:-1]}
    ON DUPLICATE KEY UPDATE bastille_register_record.work_code = VALUES(work_code);
"""

with open('assign_bastille_codes.sql', 'wt') as out:
    out.write(sql)

with open("deduplication_sheets/banned_codes.xlsx", "rb") as pth:
    banned_wbk = load_workbook(pth, read_only=True, keep_vba=False)
    banned_assigns = [
        (row[0], row[6]) for row in banned_wbk['banned_books_work_codes'].iter_rows(
            min_row=2, max_row=162, max_col=10, values_only=True
        ) if row[9] == 'Y'
    ]

sql = f"""
    INSERT INTO banned_list_record (ID, work_code)
    VALUES {str(banned_assigns)[1:-1]}
    ON DUPLICATE KEY UPDATE banned_list_record.work_code = VALUES(work_code);
"""

with open('assign_banned_codes.sql', 'wt') as out:
    out.write(sql)
