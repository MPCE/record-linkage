import mysql.connector as mariadb
from openpyxl import load_workbook

conn = mariadb.connect(user = 'root', db = 'mpce_test')
cur = conn.cursor()

agents_updated = load_workbook('agents_gendering.xlsx', read_only=True, keep_vba=False)

agents = [agent for agent in agents_updated['Sheet1'].iter_rows(min_row=2, max_row=10104, max_col=12, values_only=True)]
for i in range(len(agents)):
    agent = list(agents[i])
    agent[11] = int(agent[11])
    agents[i] = tuple(agent)

cur.execute('DROP TABLE IF EXISTS agent_temp')

cur.execute('CREATE TABLE mpce_test.`agent_temp` LIKE mpce1.agent')

cur.executemany("""
    INSERT INTO mpce_test.agent_temp (
        agent_code, name, sex, title, other_names, designation, status,
        start_date, end_date, notes, cerl_id, corporate_entity
    )
    VALUES (
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
        %s, %s
    )
""", seq_params=agents)

conn.commit()
cur.close()
conn.close()
